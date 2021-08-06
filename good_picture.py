from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os


def insert_img_to_excel(
        filename,
        by_col,
        to_col,
        img_folder
):
    """
    image to excel
    :param filename: path of file
    :param by_col: column(A, B, C ...)
    :param to_col:  insert to column (A, B, C ...)
    :param img_folder: folder of image
    :return: None
    """
    wb = load_workbook(filename)
    ws = wb.active
    for i, c in enumerate(ws[by_col], start=1):
        # print(i, c.value)
        img_ffn = os.path.join(img_folder, c.value + '.jpg')
        print(i, img_ffn)
        try:
            ws.add_image(
                img=Image(img_ffn),
                anchor=to_col + str(i)
            )
        except:
            print(c.value, 'no picture')
    wb.save(filename)


if __name__ == '__main__':
    insert_img_to_excel(
        '1.xlsx', 'A', 'B',
        img_folder="C:\\Users\\TOSHIBA\\PycharmProjects\\pythonProject1\\good_picture\\img"
    )

